from selenium import webdriver
from selenium.webdriver.common.by import By
import openpyxl

CAT_NAME = 'yalda1400' # Enter the category name of event

driver = webdriver.Chrome()

driver.get('http://ssces.ir/admin')
driver.find_element(By.ID, 'id_username').send_keys('') # Enter your username
driver.find_element(By.ID, 'id_password').send_keys('') # Enter your password
driver.find_element(By.CLASS_NAME, 'submit-row').find_element(By.TAG_NAME, 'input').click()

# click on Stu_users link
driver.find_element(By.XPATH, '//*[@id="content-main"]/div[1]/table/tbody/tr[2]/th/a').click()

# select only users that paid
driver.implicitly_wait(2)
driver.find_element(By.CLASS_NAME, 'column-IsPaid').click()
driver.implicitly_wait(2)
driver.find_element(By.CLASS_NAME, 'column-IsPaid').click()

users_link = []

def get_users_link():
    """
    Get link of each user
    """
    table = driver.find_element(By.ID, 'result_list').find_element(By.TAG_NAME, 'tbody')
    rows = table.find_elements(By.TAG_NAME, 'tr')
    for row in rows:
        cat = row.find_element(By.CLASS_NAME, 'field-cat').text
        if cat == CAT_NAME:
            user_link = row.find_element(By.CLASS_NAME, 'field-first_name').find_element(By.TAG_NAME, 'a').get_attribute('href')    
            users_link.append(user_link)

get_users_link()
driver.find_element(By.XPATH, '//*[@id="changelist-form"]/p/a[1]').click()
get_users_link()

print(f'{len(users_link)} users found')

registered_users = []
for user_link in users_link:
    """
    Crawl each user's information
    """
    driver.get(user_link)

    registered_users.append({
        'first_name': driver.find_element(By.ID, 'id_first_name').get_attribute('value'),
        'last_name': driver.find_element(By.ID, 'id_last_name').get_attribute('value'),
        'phone_number': driver.find_element(By.ID, 'id_phone').get_attribute('value'),
    })
    print(registered_users[-1])


fieldnames = ['first_name', 'last_name', 'phone_number']

# save data to excel file
wb = openpyxl.Workbook()
ws = wb.active
ws.cell(row=1, column=1).value = 'first_name'
ws.cell(row=1, column=2).value = 'last_name'
ws.cell(row=1, column=3).value = 'phone_number'

i = 2
for user in registered_users:
    ws.cell(row=i, column=1).value = user['first_name']
    ws.cell(row=i, column=2).value = user['last_name']
    ws.cell(row=i, column=3).value = user['phone_number']
    
    i += 1

wb.save('registered_users.xlsx')